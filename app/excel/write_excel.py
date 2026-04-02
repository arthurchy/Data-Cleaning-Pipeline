# write_excel.py
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.utils.logger import get_logger

logger = get_logger(__name__)

SIDE = Side(border_style="thin")
BORDER = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)
HEADER_FONT = Font(name="Arial", bold=True, size=11)
HEADER_FILL = PatternFill("solid", start_color="A9F1F5")
BODY_FONT = Font(name="Arial", size=10)
COL_FONT = Font(name="Arial", bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center")

TOP_COL_WIDTH = 40
COLNAME_WIDTH = 20
COL_WIDTH = 15

def create_worksheet(summary: dict, output_path: str, filename: str) -> None:
    """
    Write summary stats into Excel sheet
    Parameters:
        summary: summary dictionary from app.reporting.summary_stats
        output_path: excel name

    """        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dataset Profiles"
    write_data_stats(ws, summary)

    ws2 = wb.create_sheet("Column Profiles") 
    write_col_data(ws2, summary)

    ws3 = wb.create_sheet("Cleaning Log")
    write_cleaning_log(ws3, summary)
    
    os.makedirs(output_path, exist_ok=True)  
    full_path = os.path.join(output_path, filename)

    try:
        wb.save(full_path)
        logger.info(f"Excel report saved to {full_path}")
    except OSError as e:
        logger.error(f"Failed to save Excel report: {e}")
        raise

def write_header(cell, value) -> None:
    """helper function for header"""
    cell.value     = value
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER

def write_cell(cell, value) -> None:
    """helper function for body"""
    cell.value = value
    cell.font = BODY_FONT

def write_row_name(cell, value) -> None:
    """helper function for column names"""
    cell.value = value
    cell.font = COL_FONT
    cell.alignment = CENTER

def write_data_stats(ws, summary: dict) -> None:
    """
    write the dataset stats worksheet

    Parameters:
        ws: active worksheet
        summary: summary stats 
    
    """
    if not summary["dataset"]:
        logger.warning("Dataset statistic is empty")
        return

    row = 1
    DATA_HEADER = ["row_count", "column_count", "duplicate_row_removed"]
    for col_idx, label in enumerate(DATA_HEADER, start=1):
        write_header(ws.cell(row, col_idx), label)
        write_cell(ws.cell(row + 1, col_idx), summary["dataset"].get(label, "-"))

    for i in range(1, len(DATA_HEADER) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 25

def write_col_data(ws, summary: dict) -> None:
    """
    write the column stats worksheet

    Parameters:
        ws: active worksheet
        summary: summary stats 
    
    """
    if not summary["columns"]:
        logger.warning("Column statistic is empty")
        return 

    NORMAL_HEADERS = ["Column", "Type", "Missing Count", "Missing %", "Unique Count", "memory", 
                      "Top3", 
                      "Min Day", "Max Day", "Day Range",
                      "Min", "Max", "Mean", "Median", "Std Dev"
                      ]
    row = 1

    for col_idx, label in enumerate(NORMAL_HEADERS, start=1):
        write_header(ws.cell(row, col_idx), label)
        if label == "Column":
            ws.column_dimensions[get_column_letter(col_idx)].width = COLNAME_WIDTH
        elif label == "Top3":
            ws.column_dimensions[get_column_letter(col_idx)].width = TOP_COL_WIDTH
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTH

    
    for row_idx, (col_name, col_data) in enumerate(summary["columns"].items()):
        r = row_idx + 2

        write_row_name(ws.cell(r, 1), col_name)
        write_cell(ws.cell(r, 2), col_data["dtype"])
        write_cell(ws.cell(r, 3), col_data["missing_count"])
        write_cell(ws.cell(r, 4), col_data["missing_pct"])
        ws.cell(r, 4).number_format = "0.00%"

        write_cell(ws.cell(r, 5), col_data["unique_count"])
        write_cell(ws.cell(r, 6), col_data["memory"])
        
        top3 = col_data.get("Top3")
        top3_str = ("  |  ".join(f"{k}: {v}" for k, v in top3.items()) if top3 else "—")
        logger.debug(top3_str)
        write_cell(ws.cell(r, 7), top3_str)
        
        write_cell(ws.cell(r, 8), col_data.get("min_day", "-"))
        write_cell(ws.cell(r, 9), col_data.get("max_day", "-"))
        write_cell(ws.cell(r, 10), col_data.get("day_range", "-"))
        write_cell(ws.cell(r, 11), col_data.get("min", "-"))
        write_cell(ws.cell(r, 12), col_data.get("max", "-"))
        write_cell(ws.cell(r, 13), col_data.get("mean", "-"))
        write_cell(ws.cell(r, 14), col_data.get("median", "-"))
        write_cell(ws.cell(r, 15), col_data.get("std_dev", "-"))
    
    ws.freeze_panes = "B2"
    
def write_cleaning_log(ws, summary: dict) -> None:
    cleaning_log = summary.get("cleaning_log", [])
    if not cleaning_log:
        logger.warning("Cleaning log is empty")
        return

    HEADERS = ["Column", "Change Type", "Detail", "Count"]
    for col_idx, label in enumerate(HEADERS, start = 1):
        write_header(ws.cell(1, col_idx), label)
        ws.column_dimensions[get_column_letter(col_idx)].width = 30

    for row_idx, entry in enumerate(cleaning_log, start = 2):
        write_cell(ws.cell(row_idx, 1), entry["column"])
        write_cell(ws.cell(row_idx, 2), entry["change_type"])
        write_cell(ws.cell(row_idx, 3), entry["detail"])
        write_cell(ws.cell(row_idx, 4), entry["count"])

    ws.freeze_panes = "A2"


    




   
